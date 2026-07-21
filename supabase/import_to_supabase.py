# -*- coding: utf-8 -*-
"""
Load the build plan straight into Supabase (build_plan table) via the REST API.

Reads the workbook, maps the columns (same mapping as import_build_plan.py),
and upserts on seq_no so it is safe to re-run — a second run refreshes rows
rather than duplicating them. Uses only the Python standard library plus
openpyxl, and reads the project URL + anon key from ../.env (never printed).

Usage:
  python import_to_supabase.py ["path/to/SFM SCHEDULE.xlsx"]
"""
import sys, os, json, datetime, urllib.request, urllib.error
import openpyxl

HERE = os.path.dirname(__file__)
DEFAULT_XLSX = os.path.join(HERE, '..', '..', 'SFM SCHEDULE.xlsx')
ENV_PATH = os.path.join(HERE, '..', '.env')
SHEET, FIRST_DATA_ROW = 'SCHEDULE', 8
BATCH = 300

COLS = {
    'seq_no': 1, 'po_number': 32, 'part_number': 33, 'wo_number': 31, 'sage_wo': 44,
    'model': 68, 'description': 38, 'customer': 41, 'finish': 39, 'quantity': 40,
    'nest_code': 5, 'fab_cat': 8, 'sub_line': 12, 'code': 30,
    'kit_week': 6, 'fab_week': 7, 'paint_week': 9, 'paint_out_week': 10, 'subs_week': 11,
    'cut_week': 13, 'fold_week': 15, 'kitting_week': 17, 'kitting_so': 18,
    'weld_week': 20, 'painting_week': 22, 'subassembly_week': 24,
    'po_received': 42, 'original_date': 45, 'customer_req_date': 46,
    'realign_date': 48, 'delivered': 54, 'del_date': 55,
}
INT_COLS = {'seq_no', 'quantity'}


def read_env():
    env = {}
    with open(ENV_PATH, encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                k, v = line.split('=', 1)
                env[k.strip()] = v.strip().strip('"').strip("'")
    return env['VITE_SUPABASE_URL'], env['VITE_SUPABASE_ANON_KEY']


def cell(col, v):
    if v is None or v == '':
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%d/%m/%y')
    if col in INT_COLS:
        try:
            return int(v)
        except (TypeError, ValueError):
            return None
    return str(v).strip()


def build_rows(xlsx):
    ws = openpyxl.load_workbook(xlsx, data_only=True)[SHEET]
    rows, seen = [], set()
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        if (ws.cell(r, 32).value is None and ws.cell(r, 33).value is None
                and ws.cell(r, 38).value is None):
            continue
        rec = {c: cell(c, ws.cell(r, i).value) for c, i in COLS.items()}
        if rec['seq_no'] is not None:
            if rec['seq_no'] in seen:
                continue
            seen.add(rec['seq_no'])
        rows.append(rec)
    return rows


def post_batch(url, key, batch):
    endpoint = f'{url}/rest/v1/build_plan?on_conflict=seq_no'
    body = json.dumps(batch).encode('utf-8')
    req = urllib.request.Request(endpoint, data=body, method='POST', headers={
        'apikey': key,
        'Authorization': f'Bearer {key}',
        'Content-Type': 'application/json',
        'Prefer': 'resolution=merge-duplicates,return=minimal',
    })
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.status


def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    url, key = read_env()
    rows = build_rows(xlsx)
    print(f'Parsed {len(rows)} rows. Posting to {url}/rest/v1/build_plan in batches of {BATCH}...')
    total = 0
    for i in range(0, len(rows), BATCH):
        batch = rows[i:i + BATCH]
        try:
            status = post_batch(url, key, batch)
            total += len(batch)
            print(f'  batch {i // BATCH + 1}: {len(batch)} rows -> HTTP {status}  (total {total})')
        except urllib.error.HTTPError as e:
            print(f'  batch {i // BATCH + 1} FAILED: HTTP {e.code} {e.reason}')
            print('  response:', e.read().decode('utf-8', 'replace')[:800])
            sys.exit(1)
    print(f'Done. {total} rows upserted.')


if __name__ == '__main__':
    main()
