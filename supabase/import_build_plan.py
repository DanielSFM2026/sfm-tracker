# -*- coding: utf-8 -*-
"""
Build Plan importer (v1).

Reads the SFM SCHEDULE.xlsx build plan and writes build_plan.csv, matching the
build_plan table (see add_build_plan.sql). Load it via the Supabase Table Editor:
  Table Editor -> build_plan -> Insert -> Import data from CSV.

This is the manual v1 sync. The eventual automatic path reads the same workbook
from OneDrive/SharePoint via the Microsoft Graph API on a schedule and upserts
on seq_no. Re-running this against an empty table is a full refresh.

Usage:
  python import_build_plan.py "path/to/SFM SCHEDULE.xlsx"
  (defaults to ../../SFM SCHEDULE.xlsx relative to this file)
"""
import sys, csv, datetime, os
import openpyxl

DEFAULT_XLSX = os.path.join(os.path.dirname(__file__), '..', '..', 'SFM SCHEDULE.xlsx')
SHEET = 'SCHEDULE'
HEADER_ROW = 6          # column headers live on row 6
FIRST_DATA_ROW = 8

# build_plan column -> source column index (1-based) in the SCHEDULE sheet
COLS = {
    'seq_no': 1, 'po_number': 32, 'part_number': 33, 'wo_number': 31, 'sage_wo': 44,
    'model': 68, 'description': 38, 'customer': 41, 'finish': 39, 'quantity': 40,
    'nest_code': 5, 'fab_cat': 8, 'sub_line': 12, 'code': 30,
    'kit_week': 6, 'fab_week': 7, 'paint_week': 9, 'paint_out_week': 10, 'subs_week': 11,
    'cut_week': 13, 'fold_week': 15, 'kitting_week': 17, 'kitting_so': 18,
    'weld_week': 20, 'painting_week': 22, 'subassembly_week': 24,
    'po_received': 42, 'original_date': 45, 'customer_req_date': 46,
    'realign_date': 48, 'delivered': 54, 'del_date': 55,
}
INT_COLS = {'seq_no', 'quantity'}


def clean(col, v):
    if v is None:
        return ''
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%d/%m/%y')
    if col in INT_COLS:
        try:
            return str(int(v))
        except (TypeError, ValueError):
            return ''
    return str(v).strip()


def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    out = os.path.join(os.path.dirname(__file__), 'build_plan.csv')
    ws = openpyxl.load_workbook(xlsx, data_only=True)[SHEET]

    fields = list(COLS.keys())
    rows_written = 0
    seen_seq = set()
    with open(out, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in range(FIRST_DATA_ROW, ws.max_row + 1):
            # skip blank rows — a real job has at least a PO, part, or description
            if (ws.cell(r, 32).value is None and ws.cell(r, 33).value is None
                    and ws.cell(r, 38).value is None):
                continue
            rec = {col: clean(col, ws.cell(r, idx).value) for col, idx in COLS.items()}
            if rec['seq_no']:
                if rec['seq_no'] in seen_seq:
                    print(f'  ! duplicate seq_no {rec["seq_no"]} at row {r} — skipped')
                    continue
                seen_seq.add(rec['seq_no'])
            w.writerow(rec)
            rows_written += 1

    print(f'Wrote {rows_written} rows -> {out}')


if __name__ == '__main__':
    main()
